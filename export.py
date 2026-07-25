"""Excel export helpers."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re
from typing import Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_utils import value_to_display


EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def dataframe_to_excel_bytes(
    dataframe: pd.DataFrame,
    sheet_name: str = "Dados",
) -> bytes:
    """Serialize a dataframe to a formatted .xlsx file."""
    output = BytesIO()
    safe_sheet_name = _safe_sheet_name(sheet_name)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=safe_sheet_name)
        worksheet = writer.sheets[safe_sheet_name]
        _format_worksheet(worksheet, dataframe)

    output.seek(0)
    return output.getvalue()


def workbook_to_excel_bytes(sheets: Mapping[str, pd.DataFrame]) -> bytes:
    """Serialize multiple dataframes to a formatted .xlsx workbook."""
    if not sheets:
        return dataframe_to_excel_bytes(pd.DataFrame(), "Dados")

    output = BytesIO()
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for requested_name, dataframe in sheets.items():
            safe_name = _dedupe_sheet_name(
                _safe_sheet_name(requested_name),
                used_sheet_names,
            )
            dataframe.to_excel(writer, index=False, sheet_name=safe_name)
            worksheet = writer.sheets[safe_name]
            _format_worksheet(worksheet, dataframe)

    output.seek(0)
    return output.getvalue()


def timestamped_filename(prefix: str) -> str:
    """Build a stable download filename with the current local timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"


def _safe_sheet_name(sheet_name: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", sheet_name).strip()
    return (cleaned or "Dados")[:31]


def _dedupe_sheet_name(sheet_name: str, used_sheet_names: set[str]) -> str:
    candidate = sheet_name[:31]
    counter = 2
    while candidate in used_sheet_names:
        suffix = f" ({counter})"
        candidate = f"{sheet_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_sheet_names.add(candidate)
    return candidate


def _format_worksheet(worksheet, dataframe: pd.DataFrame) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937",
    )
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_index, column_name in enumerate(dataframe.columns, start=1):
        letter = get_column_letter(column_index)
        sample_values = dataframe.iloc[:500, column_index - 1].map(
            value_to_display
        )
        max_sample_width = (
            sample_values.map(len).max() if not sample_values.empty else 0
        )
        header_width = len(str(column_name))
        width = min(max(header_width, max_sample_width, 12) + 2, 60)
        worksheet.column_dimensions[letter].width = width

        for cell in worksheet[letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
